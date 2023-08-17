import argparse
import requests
from opensearchpy import OpenSearch
import numpy as np
import datetime
import time
import openpyxl
import pytz


# Notify Slack when script is done
def send_slack_notification(args, averages, excel_filename):
    if not args.webhook:
        return

    averages_text = ""
    for cache_type, average_list in averages.items():
        averages_text += f"{cache_type}: {', '.join(str(avg) for avg in average_list)}\n"

    slack_url = args.webhook
    data = {
        "value1": args.type,
        "value2": averages_text,
        "value3": excel_filename,
        "value4": args.note
    }

    response = requests.post(slack_url, json=data)

    if response.status_code == 200:
        print("Slack notification sent successfully.")
    else:
        print(f"Failed to send Slack notification: {response.status_code}")


# Expensive query to be used
def expensive_1(day, cache):
    return {
        "body": {
            "size": 0,
            "query": {
                "bool": {
                    "filter": [
                        {
                            "range": {
                                "pickup_datetime": {
                                    "gte": "2015-01-01 00:00:00",
                                    "lte": "2015-01-20 00:00:00"
                                }
                            }
                        },
                        {
                            "range": {
                                "dropoff_datetime": {
                                    "gte": "2015-01-01 00:00:00",
                                    "lte": "2015-01-20 00:00:00"
                                }
                            }
                        }
                    ]
                }
            },
            "aggs": {
                "avg_surcharge": {
                    "avg": {
                        "field": "surcharge"
                    }
                },
                "sum_total_amount": {
                    "sum": {
                        "field": "total_amount"
                    }
                }
            }
        },
        "index": 'nyc_taxis',
        "request-cache": cache,
        "request-timeout": 60
    }


# Function to send the query and measure the response time
def send_query_and_measure_time(day, endpoint, username, password, cache):
    query = expensive_1(day, cache)
    print(f"Query :  {query}")

    # Connect to the OpenSearch domain using the provided endpoint and credentials
    os = OpenSearch(
        [endpoint],
        http_auth=(username, password),
        port=443,
        use_ssl=True,
    )

    # Send the query to the OpenSearch domain
    response = os.search(index=query['index'], body=query['body'], request_timeout=60, request_cache=cache)
    took_time = response['took']

    return took_time


# Function to retrieve the cache stats to check hit counts
def get_request_cache_stats(endpoint, username, password):
    url = f"{endpoint}/_nodes/stats/indices/request_cache"
    response = requests.get(url, auth=(username, password))

    if response.status_code == 200:
        return response.json()
    else:
        print("Failed to retrieve request cache stats.")
        return None


def clearcache(args):
    # Clear cache and verify response
    url = f"{args.endpoint}/nyc_taxis/_cache/clear"
    response = requests.post(url, auth=(args.username, args.password))

    if response.status_code == 200:
        print("Request cache cleared successfully.")
    else:
        print("Failed to clear request cache." + str(response.status_code))


def process_cache_type(args, cache_type):
    # Get baseline hit count
    data = get_request_cache_stats(args.endpoint, args.username, args.password)
    hit_count = next(iter(data['nodes'].values()))['indices']['request_cache']['hit_count']

    # Get the current UTC time
    utc_datetime = datetime.datetime.now()

    # Define the PST time zone using pytz
    pst = pytz.timezone('US/Pacific')

    # Convert the UTC time to the PST time zone
    pst_datetime = utc_datetime.astimezone(pst)

    # Format the PST datetime as desired
    formatted_pst_datetime = pst_datetime.strftime("%Y-%m-%d_%H-%M-%S")

    # Create a filename using the formatted datetime
    filename = f"results_{formatted_pst_datetime}_{cache_type}_{args.note}.csv"

    num_queries = int(args.numOfQueries) # Number of times to execute the query for each date range
    save_path = 'results/'  # Path to save results

    miss_took_times = []
    daily_averages = []
    daily_p99_latencies = []
    daily_p95_latencies = []
    daily_p90_latencies = []
    daily_medians = []
    daily_mins = []
    daily_max = []

    for day in range(1, int(args.days) + 1):
        clearcache(args)  # clear cache to start
        print(f"Starting iterations for range: Jan 1 00:00:00 to Jan {day} 11:59:59")
        response_times = []
        for x in range(1, num_queries + 1):
            # time.sleep(1)
            response_time = send_query_and_measure_time(day, args.endpoint, args.username, args.password,
                                                        args.cache)  # Get took time for query
            new_hits = \
                next(iter(get_request_cache_stats(args.endpoint, args.username, args.password)['nodes'].values()))[
                    'indices']['request_cache']['hit_count']  # Check new number of hits

            if new_hits > hit_count:  # If hit count increased
                print(f"Hit. Took time: {response_time}")
                hit_count = new_hits
            else:
                miss_took_times.append(response_time)
                print(f"Miss. Took time: {response_time}")

            # Append a tuple with response time and hit/miss status
            response_times.append(response_time)
            print(f"Response {x}/{num_queries} received.")
            print(f"response_time : {response_times}")
            print(f"response_times size: {len(response_times)}")

        median = np.median(response_times[1:])
        print(f"median: {median}")
        average_response_time = sum(response_times[1:]) / (
                num_queries - 1)  # Average response time for num_queries - 1 hits, first was a miss before it got
        print(f"average_response_time : {average_response_time}")
        print(f"all response_times : {response_times}")
        print(f"sum(response_times[1:]) - 1 : {sum(response_times[1:])}")
        print(f"num_queries - 1 : {num_queries - 1}")
        # written to the cache
        p99_latency = np.percentile(response_times[1:], 99)  # Calculate p99 latency
        p95_latency = np.percentile(response_times[1:], 95)  # Calculate p95
        p90_latency = np.percentile(response_times[1:], 90)  # Calculate p90
        minimum = min(response_times[1:])
        maximum = max(response_times[1:])

        # Collect the data
        daily_averages.append((round(average_response_time, 3)))
        daily_p99_latencies.append(p99_latency)
        daily_p95_latencies.append(p95_latency)
        daily_p90_latencies.append(p90_latency)
        daily_medians.append(median)
        daily_mins.append(minimum)
        daily_max.append(maximum)

        with open(save_path + filename, 'a') as csv_file:
            csv_file.write(f'Jan 1 to Jan {str(day)} using cache of type: {cache_type} \n')
            for value in response_times:
                csv_file.write(str(value) + '\n')
            csv_file.write(f"Average response time: {average_response_time} \n")
            csv_file.write(f"Median response time: {median} \n")
            csv_file.write(f"p99 latency: {round(p99_latency, 3)} \n")
            csv_file.write(f"p95 latency: {round(p95_latency, 3)} \n")
            csv_file.write(f"p90 latency: {round(p90_latency, 3)} \n ")
            csv_file.write(f"Minimum: {min(response_times[1:])} \n ")
            csv_file.write(f"Maximum: {max(response_times[1:])} \n ")
            csv_file.write("\n")

            print(f"Results for Jan 1 to Jan {str(day)} appended to {filename}.")

    # Collect relevant data for Excel sheet
    excel_list = {
        "average_response_time": daily_averages,
        "median": daily_medians,
        "p99_latency": daily_p99_latencies,
        "p95_latency": daily_p95_latencies,
        "p90_latency": daily_p90_latencies,
        "minimum": daily_mins,
        "maximum": daily_max
    }

    # print items in tabular
    print(f"Results for cache of type {cache_type}")
    print("All average response times: ")
    for avg_time in enumerate(daily_averages, start=1):
        print(f"{avg_time}")

    print("All Miss took times: ")
    for miss_took_time in enumerate(miss_took_times, start=1):
        print(f"{miss_took_time}")

    print("All p90 response times:")
    for daily_p90_latency in enumerate(daily_p90_latencies, start=1):
        print(f"{daily_p90_latency}")

    return excel_list


def main():
    parser = argparse.ArgumentParser(description='OpenSearch Query Response Time Plotter')
    parser.add_argument('--endpoint', help='OpenSearch domain endpoint (https://example.com)')
    parser.add_argument('--username', help='Username for authentication')
    parser.add_argument('--password', help='Password for authentication')
    parser.add_argument('--days',     help='Number of days in the range to keep increasing to')
    parser.add_argument('--cache',    help='True for cache enabled and false otherwise, defaults to FALSE.', default='true')
    parser.add_argument('--type',     help='Type of cache we are using, for logging purposes', default='all')
    parser.add_argument('--webhook',  help='Slack webhook for notifying when the script is finished.', default=None)
    parser.add_argument('--numOfQueries',  help='Number of queries you want to make in each load.', default=250)
    parser.add_argument('--note',  help='Optional note to add to the test.', default="")
    args = parser.parse_args()

    caches = ['diskOnly', 'diskAndHeap', 'ehcache_heap_only', 'os_cache_only']

    if args.type != 'all':
        caches = [args.type]

    full_start_time = time.time()

    workbook = openpyxl.load_workbook('template.xlsx')  # Load Excel template
    # Get the current UTC time
    utc_datetime = datetime.datetime.now()

    # Define the PST time zone using pytz
    pst = pytz.timezone('US/Pacific')

    # Convert the UTC time to the PST time zone
    pst_datetime = utc_datetime.astimezone(pst)

    # Format the PST datetime as desired
    formatted_pst_datetime = pst_datetime.strftime("%Y-%m-%d_%H-%M-%S")
    excel_filename = f"results_{formatted_pst_datetime}_{args.type}_{args.note}.xlsx"
    workbook.save(excel_filename)

    all_averages = {}

    for cache_type in caches:  # Execute the query multiple times and measure the response time
        local_start_time = time.time()
        excel_list = process_cache_type(args, cache_type)
        local_elapsed_time = (time.time() - local_start_time) / 60
        print(f"Time taken for cache_type {cache_type} : {local_elapsed_time} minutes")
        workbook = openpyxl.load_workbook(excel_filename)  # Load Excel template
        worksheet = workbook['Sheet1']
        for index, metric in enumerate(excel_list):
            row = 4 + index * 7  # Calculate the row based on the index
            for y in range(int(args.days)):
                worksheet.cell(row=row + caches.index(cache_type), column=y + 2).value = excel_list[metric][y]
        workbook.save(excel_filename)
        all_averages[cache_type] = excel_list['average_response_time']
        time.sleep(0)  # sleep for x secs before executing the next iteration.

    full_end_time_elapsed = (time.time() - full_start_time) / 60
    print(f"Time taken for full workload : {full_end_time_elapsed} minutes")
    send_slack_notification(args, all_averages, excel_filename)


if __name__ == '__main__':
    main()
